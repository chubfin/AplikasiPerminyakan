"""Utilitas export tabel hasil ke Excel."""

from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def dataframe_to_xlsx_bytes(df: pd.DataFrame, sheet_name: str = "Hasil Perhitungan") -> bytes:
    """Mengubah DataFrame menjadi file XLSX dalam bentuk bytes."""
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=2)
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]

        max_col = len(df.columns)
        max_row = len(df) + 3

        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
        title_cell = worksheet.cell(row=1, column=1)
        title_cell.value = "Perhitungan Keekonomian Lapangan Migas"
        title_cell.font = Font(bold=True, size=14, color="FFFFFF")
        title_cell.fill = PatternFill("solid", fgColor="1F4E78")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        header_fill = PatternFill("solid", fgColor="D9EAF7")
        total_fill = PatternFill("solid", fgColor="FFF2CC")
        thin = Side(style="thin", color="808080")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for cell in worksheet[3]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        for row in worksheet.iter_rows(min_row=4, max_row=max_row, min_col=1, max_col=max_col):
            is_total = row[0].value == "Total"
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'
                if is_total:
                    cell.font = Font(bold=True)
                    cell.fill = total_fill

        for col_idx in range(1, max_col + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            for cell in worksheet[column_letter]:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))
            worksheet.column_dimensions[column_letter].width = min(max_length + 3, 24)

        worksheet.freeze_panes = "A4"
        worksheet.row_dimensions[1].height = 24
        worksheet.row_dimensions[3].height = 36

    output.seek(0)
    return output.getvalue()
